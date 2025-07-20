from tqdm import tqdm
from openpyxl import Workbook, load_workbook, styles
from utils.utils_parse import Parse
from utils.utils_datetime import DateTime
from xlrd import open_workbook


class Xlsx:
    def __init__(self) -> None:
        self.dt = DateTime()
        self.parse = Parse()
        pass

    def write_xlsx(
        self,
        filename,
        rows,
        head=[
            "车次",
            "乘车日期",
            "乘车时间",
            "发站",
            "到站",
            "车厢号",
            "座位号",
            "席别",
            "姓名",
            "证件编号",
            "证件类型",
            "性别",
            "年龄",
            "售票处",
            "窗口",
            "操作员编号",
            "售票时间",
            "票号",
            "票价",
            "票种",
        ],
    ) -> None:
        workbook: Workbook = Workbook()
        sheet = workbook.active
        sheet.append(head)
        print(f"{self.dt.get_now()} | 开始生成xlsx文件, 共有 {len(rows)} 条数据")
        for row in tqdm(rows, desc=f"写入xlsx文件", unit=f"条"):
            row: list = list(row)
            gender: str = (
                self.parse.get_gender(row[9]) if row[10] == "二代身份证" else ""
            )
            age: int | str = (
                self.parse.get_age(row[9]) if row[10] == "二代身份证" else ""
            )
            row.insert(11, gender)
            row.insert(12, age)
            sheet.append(row)
            # 如果年龄小于18岁，则设置字体颜色为红色
            if isinstance(age, int) and age < 18:
                for cell in sheet[sheet.max_row]:
                    cell.font = styles.Font(color="FF0000")  # 设置字体颜色为红色
        # 设置单元格格式为居中对齐
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = styles.Alignment(
                    horizontal="center", vertical="center"
                )
        # 设置列宽为最合适的宽度
        for index, col in enumerate(sheet.columns):
            lengths = [len(str(cell.value))
                       for cell in col if cell.value is not None]
            if lengths:
                max_length = max(lengths)
            else:
                max_length = 0
            max_length = (
                max_length + 5
                if index in [0, 1, 2, 5, 6, 7, 9, 15, 16, 17]
                else max_length * 2 + 5
            )
            sheet.column_dimensions[col[0].column_letter].width = max_length
        filename: str = filename if filename.endswith(
            f".xlsx") else f"{filename}.xlsx"
        workbook.save(filename)
        print(f"{self.dt.get_now()} | 文件 {filename} 保存完成")

    def read_xlsx(self, filename, sheet_index=0, readHead=False) -> None:
        data: list = []
        if filename.endswith(f".xlsx"):
            workbook: Workbook = load_workbook(filename=filename)
            sheet = workbook.worksheets[sheet_index]
            for row in sheet.iter_rows(values_only=True):
                if '轨迹查询' in filename:
                    data.append(self._parse_gjcx_row(row))
                elif '站站查询_' in filename or '全列查询_' in filename:
                    data.append(self._parse_zzdc_row(row))
                elif '站站查询' in filename or '全列查询' in filename:
                    data.append(self._parse_zzcx_row(row))
                elif '在逃导出数据' in filename:
                    data.append(self._parse_ztdc_row(row))
        elif filename.endswith(f".xls"):
            workbook = open_workbook(filename)
            sheet = workbook.sheet_by_index(sheet_index)
            for row_idx in range(sheet.nrows):
                row = sheet.row_values(row_idx)
                if '轨迹查询' in filename:
                    data.append(self._parse_gjcx_row(row))
                elif '站站查询_' in filename or '全列查询_' in filename:
                    data.append(self._parse_zzdc_row(row))
                elif '站站查询' in filename or '全列查询' in filename:
                    data.append(self._parse_zzcx_row(row))
                elif '在逃导出数据' in filename:
                    data.append(self._parse_ztdc_row(row))
        if not readHead:
            data = data[1:]
        return data

    def _parse_gjcx_row(self, row):
        return {
            '姓名': row[0], '证件类型': row[1], '证件编号': row[2], '乘车日期': row[3],
            '乘车时间': row[4], '车次': row[5], '发站': row[6], '到站': row[7],
            '车厢号': row[8], '席别': row[9], '座位号': row[10], '票价': row[11]
        }

    def _parse_zzcx_row(self, row):
        return {
            '姓名': row[0], '证件类型': row[1], '证件编号': row[2], '乘车日期': row[3],
            '乘车时间': row[4], '票号': row[5], '车次': row[6], '发站': row[7],
            '到站': row[8], '车厢号': row[9], '席别': row[10], '座位号': row[11],
            '票种': row[12], '票价': row[13], '售票处': row[14], '窗口': row[15],
            '操作员': row[16], '售票时间': row[17]
        }

    def _parse_ztdc_row(self, row):
        return {
            '人员编号': row[0], '姓名': row[1], '性别': row[2], '证件编号': row[3],
            '户籍地区划': row[4], '户籍地详址': row[5], '案件类别': row[6], '案件分类': row[7],
            '立案单位': row[8], '简要案情': row[9], '案件数量': row[10], '主办单位分类': row[11],
            '主办单位（区划）': row[12], '逃跑方向': row[13], '主办单位详称': row[14], '入部登记库日期': row[15]
        }

    def _parse_zzdc_row(self, row):
        return {
            '车次': row[0], '乘车日期': row[1], '乘车时间': row[2], '发站': row[3], '到站': row[4],
            '车厢号': row[5], '座位号': row[6], '席别': row[7],
            '姓名': row[8], '证件编号': row[9], '证件类型': row[10],
            '售票处': row[11], '窗口': row[12], '操作员': row[13],
            '售票时间': row[14], '票号': row[15], '票价': row[16], '票种': row[17],
        }
